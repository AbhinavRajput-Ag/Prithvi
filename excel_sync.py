import argparse
import hashlib
import json
import os
import time
from datetime import datetime
from pathlib import Path

import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / ".env"
SCHEMA_PATH = BASE_DIR / "schema.sql"
DEFAULT_WORKBOOK_PATH = BASE_DIR / "prithvi_master_sync.xlsx"

TRACKED_TABLES = [
    "farmers",
    "app_users",
    "farmer_members",
    "land_parcels",
    "crops",
    "input_costs",
    "harvest_records",
    "harvests",
    "deals",
    "buyer_registry",
    "input_suppliers",
    "soil_tests",
    "yield_estimate_revisions",
    "expense_receipts",
    "farmer_documents",
]


def load_settings():
    load_dotenv(ENV_PATH)
    return {
        "host": os.getenv("DB_HOST"),
        "database": os.getenv("DB_NAME"),
        "user": os.getenv("DB_USER"),
        "password": os.getenv("DB_PASSWORD"),
        "port": os.getenv("DB_PORT", "5432"),
        "workbook_path": Path(os.getenv("PRITHVI_EXCEL_PATH", str(DEFAULT_WORKBOOK_PATH))),
        "interval": int(os.getenv("PRITHVI_SYNC_INTERVAL", "10")),
    }


def get_connection(settings):
    return psycopg2.connect(
        host=settings["host"],
        database=settings["database"],
        user=settings["user"],
        password=settings["password"],
        port=settings["port"],
    )


def safe_mtime(path: Path):
    return path.stat().st_mtime if path.exists() else None


def hash_file(path: Path):
    if not path.exists():
        return None
    return hashlib.sha256(path.read_bytes()).hexdigest()


def fetch_table_columns(cursor, table_name):
    cursor.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
        ORDER BY ordinal_position;
        """,
        (table_name,),
    )
    return [row[0] for row in cursor.fetchall()]


def fetch_table_rows(cursor, table_name, columns):
    if not columns:
        return []
    quoted_columns = ", ".join(f'"{column}"' for column in columns)
    cursor.execute(f'SELECT {quoted_columns} FROM "{table_name}" ORDER BY 1;')
    return cursor.fetchall()


def compute_db_snapshot(settings):
    snapshot = {"tables": {}, "schema_fields": [], "field_coverage": []}
    with get_connection(settings) as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT table_name, column_name, data_type, is_nullable, column_default
                FROM information_schema.columns
                WHERE table_schema = 'public' AND table_name = ANY(%s)
                ORDER BY table_name, ordinal_position;
                """,
                (TRACKED_TABLES,),
            )
            snapshot["schema_fields"] = cursor.fetchall()

            for table_name in TRACKED_TABLES:
                columns = fetch_table_columns(cursor, table_name)
                rows = fetch_table_rows(cursor, table_name, columns)
                serialized_rows = json.dumps(rows, default=str, ensure_ascii=True, sort_keys=False)
                fingerprint = hashlib.sha256(serialized_rows.encode("utf-8")).hexdigest()
                snapshot["tables"][table_name] = {
                    "columns": columns,
                    "rows": rows,
                    "row_count": len(rows),
                    "fingerprint": fingerprint,
                }
                for column in columns:
                    cursor.execute(
                        f'SELECT COUNT(*) FROM "{table_name}" WHERE "{column}" IS NOT NULL AND CAST("{column}" AS TEXT) <> \'\';'
                    )
                    populated_count = cursor.fetchone()[0]
                    snapshot["field_coverage"].append(
                        (table_name, column, len(rows), populated_count)
                    )
    return snapshot


def state_signature(snapshot, schema_hash, workbook_mtime):
    compact = {
        "schema_hash": schema_hash,
        "workbook_mtime": workbook_mtime,
        "tables": {
            table_name: {
                "row_count": meta["row_count"],
                "fingerprint": meta["fingerprint"],
            }
            for table_name, meta in snapshot["tables"].items()
        },
    }
    return hashlib.sha256(json.dumps(compact, sort_keys=True).encode("utf-8")).hexdigest()


def autosize_sheet(sheet):
    for column_cells in sheet.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 42)


def append_sheet_rows(sheet, headers, rows):
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(list(row))
    autosize_sheet(sheet)


def write_workbook(settings, snapshot):
    workbook_path = settings["workbook_path"]
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    status_sheet = workbook.active
    status_sheet.title = "SyncStatus"
    append_sheet_rows(
        status_sheet,
        ["key", "value"],
        [
            ("generated_at", datetime.now().isoformat(timespec="seconds")),
            ("database", settings["database"]),
            ("host", settings["host"]),
            ("schema_path", str(SCHEMA_PATH)),
            ("workbook_path", str(workbook_path)),
            ("tracked_tables", ", ".join(TRACKED_TABLES)),
        ],
    )

    summary_sheet = workbook.create_sheet("TableSummary")
    append_sheet_rows(
        summary_sheet,
        ["table_name", "row_count", "fingerprint"],
        [
            (table_name, meta["row_count"], meta["fingerprint"])
            for table_name, meta in snapshot["tables"].items()
        ],
    )

    schema_sheet = workbook.create_sheet("SchemaFields")
    append_sheet_rows(
        schema_sheet,
        ["table_name", "column_name", "data_type", "is_nullable", "column_default"],
        snapshot["schema_fields"],
    )

    coverage_sheet = workbook.create_sheet("FieldCoverage")
    append_sheet_rows(
        coverage_sheet,
        ["table_name", "column_name", "row_count", "populated_count", "empty_count"],
        [
            (table_name, column_name, row_count, populated_count, row_count - populated_count)
            for table_name, column_name, row_count, populated_count in snapshot["field_coverage"]
        ],
    )

    for table_name, meta in snapshot["tables"].items():
        sheet = workbook.create_sheet(table_name[:31])
        append_sheet_rows(sheet, meta["columns"], meta["rows"])

    workbook.save(workbook_path)


def sync_once(settings):
    snapshot = compute_db_snapshot(settings)
    write_workbook(settings, snapshot)
    return snapshot


def watch_sync(settings):
    last_signature = None
    workbook_path = settings["workbook_path"]

    while True:
        schema_hash = hash_file(SCHEMA_PATH)
        workbook_mtime = safe_mtime(workbook_path)
        snapshot = compute_db_snapshot(settings)
        current_signature = state_signature(snapshot, schema_hash, workbook_mtime)

        if current_signature != last_signature:
            write_workbook(settings, snapshot)
            last_signature = state_signature(snapshot, schema_hash, safe_mtime(workbook_path))
            print(f"[{datetime.now().isoformat(timespec='seconds')}] workbook synced")
        else:
            print(f"[{datetime.now().isoformat(timespec='seconds')}] no change detected")

        time.sleep(settings["interval"])


def main():
    parser = argparse.ArgumentParser(description="Sync PostgreSQL + schema data into a master Excel workbook.")
    parser.add_argument("--once", action="store_true", help="Run one sync and exit.")
    parser.add_argument("--interval", type=int, default=None, help="Polling interval in seconds for watch mode.")
    args = parser.parse_args()

    settings = load_settings()
    if args.interval is not None:
        settings["interval"] = args.interval

    if args.once:
        sync_once(settings)
        print(f"Workbook written to {settings['workbook_path']}")
        return

    watch_sync(settings)


if __name__ == "__main__":
    main()
